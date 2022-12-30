import csv
import re
from collections import Counter
import openpyxl
from openpyxl.styles import Border, Side, Font
import matplotlib.pyplot as plt
import numpy as np
import os.path
from jinja2 import Environment, FileSystemLoader
import pdfkit




def main(prof, stats):
    def to_string(ws):
        result = ""
        result += "<table>\n"

        for row in ws.rows:
            result += "<tr>"
            for value in row:
                if value.value == None:
                    result += "<th style=\"border: 0;\"></th>"
                    continue
                result += f"<th>{value.value}</th>"
            result += "</tr>"

        result += "</table>"
        return result

    class report:
        def __init__(self):
            self.file = openpyxl.Workbook()

            ws = self.file.active
            ws.title = "Статистика по годам"
            self.file.create_sheet("Статистика по городам")

        def generate_pdf(self, prof):
            ws = to_string(self.file["Статистика по годам"])
            env = Environment(loader=FileSystemLoader('.'))
            template = env.get_template("pdf_template.html")
            pdf_template = template.render({'prof': prof, 'ws1': ws})
            config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltopdf\bin\wkhtmltopdf.exe')
            pdfkit.from_string(pdf_template, 'report.pdf', configuration=config,
                               options={"enable-local-file-access": ""})

        def generate_excel(self, list_years):
            ws = self.file["Статистика по годам"]
            ws.append(["Год", "Средняя зарплата", f"Средняя зарплата - {prof}", "Количество вакансий",
                       f"Количество вакансий - {prof}"])
            ws2 = self.file["Статистика по городам"]
            years = list(list_years[0].keys())
            l = len(years)
            i = 0
            flag = True
            thins = Side(border_style="thin", color="000000")
            for col in ws.iter_cols(min_row=2, max_col=5, max_row=l + 1):
                min = 2003
                if flag:
                    j = 0
                    for cell in col:
                        cell.value = years[j]
                        cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
                        j += 1
                    flag = False
                else:
                    for cell in col:
                        cell.value = list_years[i][min]
                        cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
                        min += 1
                    i += 1
            for col in ws.columns:
                length = max(len(str(cell.value)) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = length + 2
            ws2["A1"].value = "Город"
            ws2["A1"].font = Font(bold=True)
            ws2["B1"].value = "Уровень зарплаты"
            ws2["B1"].font = Font(bold=True)
            ws2["D1"].value = "Город"
            ws2["D1"].font = Font(bold=True)
            ws2["E1"].value = "Доля вакансий"
            ws2["E1"].font = Font(bold=True)
            self.file.save('report.xlsx')


    report = report()
    report.generate_excel(stats)
    report.generate_pdf(prof)
